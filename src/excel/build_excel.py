import os
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from openpyxl.chart import BarChart, Reference
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule

wb = Workbook()

# ── COLOURS ──
PURPLE_DARK  = '3B1FA8'
PURPLE_MID   = '5B21B6'
PURPLE_LIGHT = 'EDE9FE'
CYAN_DARK    = '0E7490'
CYAN_LIGHT   = 'E0F7FA'
DARK_BG      = '1E1B4B'
DARK_TEXT    = '1F2937'
GRAY_TEXT    = '6B7280'
WHITE        = 'FFFFFF'
LIGHT_ROW    = 'F8F7FF'
LIGHT_ROW2   = 'FFFFFF'
SUCCESS      = '166534'
SUCCESS_BG   = 'DCFCE7'
WARN         = '92400E'
WARN_BG      = 'FEF3C7'
DANGER       = '991B1B'
DANGER_BG    = 'FEE2E2'

def font(bold=False, color=DARK_TEXT, size=10, italic=False):
    return Font(name='Arial', bold=bold, color=color, size=size, italic=italic)

def fill(color):
    return PatternFill('solid', fgColor=color)

def border(style='thin', color='D1D5DB'):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def thin_border():
    return border('thin', 'D1D5DB')

def center():
    return Alignment(horizontal='center', vertical='center', wrap_text=True)

def left():
    return Alignment(horizontal='left', vertical='center', wrap_text=True)

def right_align():
    return Alignment(horizontal='right', vertical='center')

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def header_row(ws, row, values, widths=None, bg=PURPLE_MID, fg=WHITE, size=10):
    for i, val in enumerate(values, 1):
        c = ws.cell(row=row, column=i, value=val)
        c.font = font(bold=True, color=fg, size=size)
        c.fill = fill(bg)
        c.alignment = center()
        c.border = thin_border()
    ws.row_dimensions[row].height = 28

def data_row(ws, row, values, formats=None, bolds=None, colors=None, bg=None):
    row_bg = bg or (LIGHT_ROW if row % 2 == 0 else LIGHT_ROW2)
    for i, val in enumerate(values, 1):
        c = ws.cell(row=row, column=i, value=val)
        c.font = font(bold=bolds[i-1] if bolds else False,
                      color=colors[i-1] if colors else DARK_TEXT)
        c.fill = fill(row_bg)
        c.alignment = center() if i > 1 else left()
        c.border = thin_border()
        if formats and formats[i-1]:
            c.number_format = formats[i-1]
    ws.row_dimensions[row].height = 22

def title_cell(ws, row, col, text, colspan=1, bg=DARK_BG, size=14):
    c = ws.cell(row=row, column=col, value=text)
    c.font = font(bold=True, color=WHITE, size=size)
    c.fill = fill(bg)
    c.alignment = center()
    if colspan > 1:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+colspan-1)
    ws.row_dimensions[row].height = 36

def subtitle_cell(ws, row, col, text, colspan=1, bg=PURPLE_LIGHT, size=10, color=PURPLE_MID):
    c = ws.cell(row=row, column=col, value=text)
    c.font = font(bold=True, color=color, size=size)
    c.fill = fill(bg)
    c.alignment = left()
    if colspan > 1:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+colspan-1)
    ws.row_dimensions[row].height = 22

def spacer(ws, row):
    for col in range(1, 10):
        c = ws.cell(row=row, column=col, value='')
        c.fill = fill(WHITE)
    ws.row_dimensions[row].height = 10

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 1: OVERVIEW
# ═══════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = 'Overview'
ws1.sheet_view.showGridLines = False

# Column widths
for col, w in [(1,26),(2,14),(3,14),(4,14),(5,14),(6,14),(7,14),(8,14)]:
    set_col_width(ws1, col, w)

title_cell(ws1, 1, 1, 'ShieldVIN — Vehicle Identity on Midnight Network', 8, DARK_BG, 14)
subtitle_cell(ws1, 2, 1, 'Financial Model Overview  |  2026–2030  |  All figures in USD unless stated', 8, PURPLE_LIGHT, 10, PURPLE_MID)
spacer(ws1, 3)

# ── KEY METRICS ──
subtitle_cell(ws1, 4, 1, '  KEY MARKET METRICS', 4, DARK_BG, 10, WHITE)
for col in range(5, 9):
    ws1.cell(row=4, column=col).fill = fill(DARK_BG)
ws1.row_dimensions[4].height = 22

metrics = [
    ('Global Annual Vehicle Theft Losses (USD)', '$21.8B', 'Source: FBI / INTERPOL / SAPS estimates'),
    ('US Annual Losses Alone', '$7.4B', 'Source: FBI UCR 2023'),
    ('Global New Vehicle Production / Year', '~80M units', 'Source: OICA 2024'),
    ('VIN Cloning: Estimated % of Stolen Vehicles Resold', '~35%', 'Estimated — varies by market'),
    ('Insurance Cost of VIN Fraud (Global, annual)', '$4.2B', 'Estimated from industry reports'),
    ('Average Premium Uplift (VIN fraud risk)', '$180–$320/yr', 'Per vehicle in high-theft markets'),
]
header_row(ws1, 5, ['Metric', 'Value', 'Note / Source', '', ''], bg=PURPLE_MID)
ws1.merge_cells('C5:E5')
for i, (metric, val, note) in enumerate(metrics):
    r = i + 6
    ws1.cell(r, 1, metric).font = font(bold=False, color=DARK_TEXT)
    ws1.cell(r, 1).fill = fill(LIGHT_ROW if i % 2 == 0 else LIGHT_ROW2)
    ws1.cell(r, 1).border = thin_border()
    ws1.cell(r, 1).alignment = left()
    ws1.cell(r, 2, val).font = font(bold=True, color=PURPLE_MID)
    ws1.cell(r, 2).fill = fill(LIGHT_ROW if i % 2 == 0 else LIGHT_ROW2)
    ws1.cell(r, 2).border = thin_border()
    ws1.cell(r, 2).alignment = center()
    ws1.cell(r, 3, note).font = font(color=GRAY_TEXT, italic=True)
    ws1.merge_cells(f'C{r}:E{r}')
    ws1.cell(r, 3).fill = fill(LIGHT_ROW if i % 2 == 0 else LIGHT_ROW2)
    ws1.cell(r, 3).border = thin_border()
    ws1.cell(r, 3).alignment = left()
    ws1.row_dimensions[r].height = 20

spacer(ws1, 13)

# ── REVENUE SUMMARY ──
subtitle_cell(ws1, 14, 1, '  5-YEAR REVENUE SUMMARY  (USD Millions)', 7, DARK_BG, 10, WHITE)
ws1.row_dimensions[14].height = 22
header_row(ws1, 15, ['Revenue Stream', 'Year 1\n2026', 'Year 2\n2027', 'Year 3\n2028', 'Year 4\n2029', 'Year 5\n2030', 'CAGR', 'Notes'], bg=PURPLE_MID)

rev_streams = [
    ('Vehicle Identity Token Minting', 0.10, 0.80, 4.20, 18.50, 58.00, 'Per-vehicle fee at manufacture. Volume tiers $8–$25/vehicle.'),
    ('Verification API Access', 0.05, 0.30, 1.80, 9.20, 31.00, 'Insurance, dealers, fleet operators. $0.10–$2.50/query.'),
    ('Insurance & Enterprise Partnerships', 0.00, 0.10, 0.90, 5.40, 19.00, 'Shared claim savings + data products.'),
    ('Total Revenue', 0.15, 1.20, 6.90, 33.10, 108.00, ''),
]
for i, (stream, y1, y2, y3, y4, y5, note) in enumerate(rev_streams):
    r = i + 16
    is_total = stream == 'Total Revenue'
    bg_color = DARK_BG if is_total else (LIGHT_ROW if i % 2 == 0 else LIGHT_ROW2)
    fg_color = WHITE if is_total else DARK_TEXT
    
    ws1.cell(r, 1, stream).font = font(bold=is_total, color=fg_color)
    ws1.cell(r, 1).fill = fill(bg_color)
    ws1.cell(r, 1).border = thin_border()
    ws1.cell(r, 1).alignment = left()
    
    if not is_total:
        cagr_formula = f'=(F{r}/B{r})^(1/4)-1'
    else:
        cagr_formula = f'=(F{r}/B{r})^(1/4)-1'
    
    for j, val in enumerate([y1, y2, y3, y4, y5], 2):
        c = ws1.cell(r, j, val)
        c.font = font(bold=is_total, color=WHITE if is_total else PURPLE_MID)
        c.fill = fill(bg_color)
        c.border = thin_border()
        c.alignment = center()
        c.number_format = '$#,##0.00"M"'
    
    if not is_total:
        cagr_cell = ws1.cell(r, 7, f'=(F{r}/B{r})^(1/4)-1')
    else:
        cagr_cell = ws1.cell(r, 7, f'=(F{r}/B{r})^(1/4)-1')
    cagr_cell.font = font(bold=is_total, color=WHITE if is_total else CYAN_DARK)
    cagr_cell.fill = fill(bg_color)
    cagr_cell.border = thin_border()
    cagr_cell.alignment = center()
    cagr_cell.number_format = '0.0%'
    
    ws1.cell(r, 8, note).font = font(color=GRAY_TEXT if not is_total else WHITE, italic=not is_total)
    ws1.cell(r, 8).fill = fill(bg_color)
    ws1.cell(r, 8).border = thin_border()
    ws1.cell(r, 8).alignment = left()
    ws1.row_dimensions[r].height = 22

spacer(ws1, 21)

# ── INVESTMENT REQUIRED ──
subtitle_cell(ws1, 22, 1, '  INVESTMENT REQUIREMENTS', 4, DARK_BG, 10, WHITE)
for col in range(5, 9):
    ws1.cell(row=22, column=col).fill = fill(DARK_BG)

header_row(ws1, 23, ['Phase', 'Period', 'Investment Required', 'Key Milestones', ''], bg=PURPLE_MID)
ws1.merge_cells('D23:E23')

investments = [
    ('Phase 1 — PoC & Consortium', 'Months 0–12', '$3M – $5M', 'Midnight testnet smart contract, key ceremony PoC, founding OEM + insurer MoU'),
    ('Phase 2 — Pilot Programme', 'Months 12–30', '$9M – $13M', 'Live vehicles, law enforcement integration, commercial API launch'),
    ('Phase 1+2 Total', '', '$12M – $18M', 'Target: self-sustaining by Year 3 on API revenue'),
    ('Phase 3 — Rollout', 'Months 30–54', '$25M – $40M', '500K+ vehicles, 3+ jurisdictions, multi-manufacturer'),
    ('Phase 4 — Global Standard', 'Months 54+', 'Revenue-funded', 'Regulatory mandate lobbying, 5M+ vehicles, ISO/UNECE recognition'),
]

for i, (phase, period, invest, milestone) in enumerate(investments):
    r = i + 24
    is_total = 'Total' in phase
    bg_color = CYAN_LIGHT if is_total else (LIGHT_ROW if i % 2 == 0 else LIGHT_ROW2)
    
    ws1.cell(r, 1, phase).font = font(bold=is_total, color=DARK_TEXT)
    ws1.cell(r, 1).fill = fill(bg_color)
    ws1.cell(r, 1).border = thin_border()
    ws1.cell(r, 1).alignment = left()
    
    ws1.cell(r, 2, period).font = font(color=GRAY_TEXT)
    ws1.cell(r, 2).fill = fill(bg_color)
    ws1.cell(r, 2).border = thin_border()
    ws1.cell(r, 2).alignment = center()
    
    ws1.cell(r, 3, invest).font = font(bold=True, color=CYAN_DARK)
    ws1.cell(r, 3).fill = fill(bg_color)
    ws1.cell(r, 3).border = thin_border()
    ws1.cell(r, 3).alignment = center()
    
    ws1.merge_cells(f'D{r}:E{r}')
    ws1.cell(r, 4, milestone).font = font(color=DARK_TEXT, italic=is_total)
    ws1.cell(r, 4).fill = fill(bg_color)
    ws1.cell(r, 4).border = thin_border()
    ws1.cell(r, 4).alignment = left()
    ws1.row_dimensions[r].height = 22


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 2: DETAILED REVENUE MODEL
# ═══════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet('Revenue Model')
ws2.sheet_view.showGridLines = False

for col, w in [(1,30),(2,16),(3,16),(4,16),(5,16),(6,16),(7,16)]:
    set_col_width(ws2, col, w)

title_cell(ws2, 1, 1, 'Detailed Revenue Model — Assumptions & Projections', 7, DARK_BG, 14)
subtitle_cell(ws2, 2, 1, 'Blue text = input assumptions (change these to model scenarios)  |  Black text = calculated outputs', 7, PURPLE_LIGHT, 9, PURPLE_MID)
spacer(ws2, 3)

# ── MINTING ASSUMPTIONS ──
subtitle_cell(ws2, 4, 1, '  STREAM 1: VEHICLE IDENTITY TOKEN MINTING', 4, PURPLE_MID, 10, WHITE)
for col in range(5, 8): ws2.cell(4, col).fill = fill(PURPLE_MID)

header_row(ws2, 5, ['Assumption', 'Year 1', 'Year 2', 'Year 3', 'Year 4', 'Year 5', 'Unit'], bg=PURPLE_MID)

minting_inputs = [
    ('Global new vehicle production', 80_000_000, 82_000_000, 84_000_000, 86_000_000, 88_000_000, 'Units/year'),
    ('% of production enrolled in system', 0.001, 0.005, 0.02, 0.08, 0.25, '% (input)'),
    ('Vehicles enrolled (calculated)', '=B7*B8', '=C7*C8', '=D7*D8', '=E7*E8', '=F7*F8', 'Units'),
    ('Average fee per token minted ($)', 12, 12, 11, 10, 9, 'USD (input)'),
    ('Minting Revenue', '=B9*B10/1000000', '=C9*C10/1000000', '=D9*D10/1000000', '=E9*E10/1000000', '=F9*F10/1000000', '$M'),
]

for i, (label, y1, y2, y3, y4, y5, unit) in enumerate(minting_inputs):
    r = i + 6
    is_calc = not isinstance(y1, (int, float))
    is_revenue = label == 'Minting Revenue'
    bg = SUCCESS_BG if is_revenue else (LIGHT_ROW if i % 2 == 0 else LIGHT_ROW2)
    fg_val = SUCCESS if is_revenue else (DARK_TEXT if is_calc else '0000CC')
    
    ws2.cell(r, 1, label).font = font(bold=is_revenue, color=DARK_TEXT)
    ws2.cell(r, 1).fill = fill(bg)
    ws2.cell(r, 1).border = thin_border()
    ws2.cell(r, 1).alignment = left()
    
    for j, val in enumerate([y1, y2, y3, y4, y5], 2):
        c = ws2.cell(r, j, val)
        c.font = font(bold=is_revenue, color=fg_val)
        c.fill = fill(bg)
        c.border = thin_border()
        c.alignment = center()
        if label == '% of production enrolled in system':
            c.number_format = '0.00%'
        elif label == 'Minting Revenue':
            c.number_format = '$#,##0.00"M"'
        elif 'enrolled' in label and is_calc:
            c.number_format = '#,##0'
        elif isinstance(val, (int, float)) and val > 1000:
            c.number_format = '#,##0'
    
    ws2.cell(r, 7, unit).font = font(color=GRAY_TEXT, italic=True)
    ws2.cell(r, 7).fill = fill(bg)
    ws2.cell(r, 7).border = thin_border()
    ws2.cell(r, 7).alignment = left()
    ws2.row_dimensions[r].height = 22

spacer(ws2, 12)

# ── API REVENUE ASSUMPTIONS ──
subtitle_cell(ws2, 13, 1, '  STREAM 2: VERIFICATION API ACCESS', 4, CYAN_DARK, 10, WHITE)
for col in range(5, 8): ws2.cell(13, col).fill = fill(CYAN_DARK)

header_row(ws2, 14, ['Assumption', 'Year 1', 'Year 2', 'Year 3', 'Year 4', 'Year 5', 'Unit'], bg=CYAN_DARK)

api_inputs = [
    ('Active enrolled vehicles (cumulative)', '=B9', '=B9+C9', '=B15+D9', '=C15+E9', '=D15+F9', 'Units'),
    ('Avg verifications per vehicle per year', 4, 5, 6, 7, 8, 'Queries/vehicle/yr'),
    ('Total verification queries (M)', '=B15*B16/1000000', '=C15*C16/1000000', '=D15*D16/1000000', '=E15*E16/1000000', '=F15*F16/1000000', 'Millions'),
    ('% queries paid (excl. govt/police)', 0.60, 0.62, 0.65, 0.68, 0.70, '% (input)'),
    ('Average price per paid query ($)', 0.18, 0.20, 0.22, 0.25, 0.28, 'USD (input)'),
    ('API Revenue', '=B17*B18*B19/1', '=C17*C18*C19/1', '=D17*D18*D19/1', '=E17*E18*E19/1', '=F17*F18*F19/1', '$M'),
]

for i, (label, y1, y2, y3, y4, y5, unit) in enumerate(api_inputs):
    r = i + 15
    is_calc = not isinstance(y1, (int, float))
    is_revenue = label == 'API Revenue'
    bg = SUCCESS_BG if is_revenue else (LIGHT_ROW if i % 2 == 0 else LIGHT_ROW2)
    fg_val = SUCCESS if is_revenue else (DARK_TEXT if is_calc else '0000CC')
    
    ws2.cell(r, 1, label).font = font(bold=is_revenue, color=DARK_TEXT)
    ws2.cell(r, 1).fill = fill(bg)
    ws2.cell(r, 1).border = thin_border()
    ws2.cell(r, 1).alignment = left()
    
    for j, val in enumerate([y1, y2, y3, y4, y5], 2):
        c = ws2.cell(r, j, val)
        c.font = font(bold=is_revenue, color=fg_val)
        c.fill = fill(bg)
        c.border = thin_border()
        c.alignment = center()
        if '%' in label:
            c.number_format = '0%'
        elif is_revenue:
            c.number_format = '$#,##0.00"M"'
        elif 'queries' in label.lower() and is_calc:
            c.number_format = '#,##0.0"M"'
    
    ws2.cell(r, 7, unit).font = font(color=GRAY_TEXT, italic=True)
    ws2.cell(r, 7).fill = fill(bg)
    ws2.cell(r, 7).border = thin_border()
    ws2.cell(r, 7).alignment = left()
    ws2.row_dimensions[r].height = 22

spacer(ws2, 22)

# ── ENTERPRISE REVENUE ──
subtitle_cell(ws2, 23, 1, '  STREAM 3: INSURANCE & ENTERPRISE PARTNERSHIPS', 4, PURPLE_MID, 10, WHITE)
for col in range(5, 8): ws2.cell(23, col).fill = fill(PURPLE_MID)

header_row(ws2, 24, ['Assumption', 'Year 1', 'Year 2', 'Year 3', 'Year 4', 'Year 5', 'Unit'], bg=PURPLE_MID)

enterprise_data = [
    ('Major insurer partners', 0, 1, 2, 4, 7, 'Count'),
    ('Annual contract value per insurer ($M)', 0, 0.10, 0.25, 0.75, 1.50, '$M (input)'),
    ('Insurance partnership revenue', '=B25*B26', '=C25*C26', '=D25*D26', '=E25*E26', '=F25*F26', '$M'),
    ('Enterprise data product revenue', 0.00, 0.00, 0.25, 1.15, 4.50, '$M (input)'),
    ('Enterprise & Insurance Total', '=B27+B28', '=C27+C28', '=D27+D28', '=E27+E28', '=F27+F28', '$M'),
]

for i, (label, y1, y2, y3, y4, y5, unit) in enumerate(enterprise_data):
    r = i + 25
    is_calc = not isinstance(y1, (int, float))
    is_total = 'Total' in label
    bg = SUCCESS_BG if is_total else (LIGHT_ROW if i % 2 == 0 else LIGHT_ROW2)
    fg_val = SUCCESS if is_total else (DARK_TEXT if is_calc else '0000CC')
    
    ws2.cell(r, 1, label).font = font(bold=is_total, color=DARK_TEXT)
    ws2.cell(r, 1).fill = fill(bg)
    ws2.cell(r, 1).border = thin_border()
    ws2.cell(r, 1).alignment = left()
    
    for j, val in enumerate([y1, y2, y3, y4, y5], 2):
        c = ws2.cell(r, j, val)
        c.font = font(bold=is_total, color=fg_val)
        c.fill = fill(bg)
        c.border = thin_border()
        c.alignment = center()
        if is_total or 'revenue' in label.lower() or '$M' in unit:
            c.number_format = '$#,##0.00"M"'
    
    ws2.cell(r, 7, unit).font = font(color=GRAY_TEXT, italic=True)
    ws2.cell(r, 7).fill = fill(bg)
    ws2.cell(r, 7).border = thin_border()
    ws2.cell(r, 7).alignment = left()
    ws2.row_dimensions[r].height = 22

spacer(ws2, 31)

# ── TOTAL SUMMARY ──
subtitle_cell(ws2, 32, 1, '  CONSOLIDATED REVENUE SUMMARY', 4, DARK_BG, 10, WHITE)
for col in range(5, 8): ws2.cell(32, col).fill = fill(DARK_BG)
header_row(ws2, 33, ['Revenue Stream', 'Year 1', 'Year 2', 'Year 3', 'Year 4', 'Year 5', 'Notes'], bg=DARK_BG)

summary_rows = [
    ('Minting Revenue ($M)', '=B11', '=C11', '=D11', '=E11', '=F11', 'From Stream 1 above'),
    ('API Verification Revenue ($M)', '=B20', '=C20', '=D20', '=E20', '=F20', 'From Stream 2 above'),
    ('Insurance & Enterprise ($M)', '=B29', '=C29', '=D29', '=E29', '=F29', 'From Stream 3 above'),
    ('TOTAL REVENUE ($M)', '=B34+B35+B36', '=C34+C35+C36', '=D34+D35+D36', '=E34+E35+E36', '=F34+F35+F36', 'Consolidated'),
]

for i, (label, y1, y2, y3, y4, y5, note) in enumerate(summary_rows):
    r = i + 34
    is_total = 'TOTAL' in label
    bg = DARK_BG if is_total else (LIGHT_ROW if i % 2 == 0 else LIGHT_ROW2)
    fg = WHITE if is_total else DARK_TEXT
    
    ws2.cell(r, 1, label).font = font(bold=is_total, color=fg)
    ws2.cell(r, 1).fill = fill(bg)
    ws2.cell(r, 1).border = thin_border()
    ws2.cell(r, 1).alignment = left()
    
    for j, val in enumerate([y1, y2, y3, y4, y5], 2):
        c = ws2.cell(r, j, val)
        c.font = font(bold=is_total, color=WHITE if is_total else PURPLE_MID)
        c.fill = fill(bg)
        c.border = thin_border()
        c.alignment = center()
        c.number_format = '$#,##0.00"M"'
    
    ws2.cell(r, 7, note).font = font(color=WHITE if is_total else GRAY_TEXT, italic=True)
    ws2.cell(r, 7).fill = fill(bg)
    ws2.cell(r, 7).border = thin_border()
    ws2.cell(r, 7).alignment = left()
    ws2.row_dimensions[r].height = 22 if not is_total else 28


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 3: RISK MATRIX
# ═══════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet('Risk Matrix')
ws3.sheet_view.showGridLines = False

for col, w in [(1,4),(2,30),(3,14),(4,14),(5,10),(6,34)]:
    set_col_width(ws3, col, w)

title_cell(ws3, 1, 1, 'Risk Register — ShieldVIN Vehicle Identity System', 6, DARK_BG, 14)
subtitle_cell(ws3, 2, 1, 'Likelihood and Impact scored 1–5. Risk Score = Likelihood × Impact. Colour coded by score.', 6, PURPLE_LIGHT, 9, PURPLE_MID)
spacer(ws3, 3)

header_row(ws3, 4, ['#', 'Risk Description', 'Category', 'Likelihood\n(1–5)', 'Impact\n(1–5)', 'Mitigation Strategy'], bg=PURPLE_MID)

risks = [
    # Technical
    ('T1', 'Secure Element chip physically attacked and key extracted', 'Technical', 2, 5, 'Three-node requirement means one chip compromise is insufficient. VSE-1 spec mandates side-channel resistance and regular security certification.'),
    ('T2', 'Man-in-the-middle attack on CAN bus inter-node communication', 'Technical', 3, 3, 'All inter-node communication is encrypted with session-unique challenge-response nonces. Replay attacks blocked by design.'),
    ('T3', 'Midnight Network outage or validator failure', 'Technical', 2, 3, 'Offline status certificates valid for 24 hours. Decentralised validator set eliminates single point of failure.'),
    ('T4', 'Quantum computing attack on elliptic curve cryptography', 'Technical', 1, 5, 'VSE-1 crypto-agility requirement mandates post-quantum upgrade path. Governance body monitors and mandates migration timeline.'),
    ('T5', 'Key generation ceremony compromised at manufacture', 'Technical', 2, 5, 'Strictly audited ceremony process. Hardware security module oversight. Tamper-evident ceremony logging. Multiple independent witnesses required.'),
    # Commercial
    ('C1', 'Manufacturers refuse to agree on a common standard', 'Commercial', 3, 5, 'Engage ISO/SAE as neutral formal custodians. Start with willing early adopters. Insurance and regulatory pressure accelerates holdouts.'),
    ('C2', 'Consumer resistance to perceived tracking or surveillance', 'Commercial', 3, 3, 'Clear privacy messaging — no location data stored. ZK proofs mean queries reveal nothing. Owner controls all access.'),
    ('C3', 'Chip integration cost unacceptable in price-sensitive segments', 'Commercial', 3, 2, 'At scale, SE chips add ~$15–$40 per vehicle. Insurance discounts of $50–$200/year offset cost for consumers. OEM feature differentiation value.'),
    ('C4', 'Rival blockchain consortium launches competing solution first', 'Commercial', 2, 3, 'First-mover on standards is decisive. VSE-1/VIT-1 adoption creates massive switching costs. Speed of Phase 1 is critical.'),
    ('C5', 'Insurance industry does not engage as partners', 'Commercial', 2, 4, 'Financial case is clear — direct claim cost reduction. Approach via Lloyd\'s / major reinsurers as early adopters who can pull retail insurers.'),
    # Regulatory
    ('R1', 'Regulators reject blockchain-based identity as legally valid', 'Regulatory', 3, 5, 'Position as supplement to existing systems initially. Build API outputs compatible with existing government formats. Engage regulators in pilot phase design.'),
    ('R2', 'GDPR or privacy law non-compliance found in design', 'Regulatory', 2, 4, 'ZK architecture means no personal data on-chain by design. Engage specialist privacy law firm pre-launch for architecture validation.'),
    ('R3', 'Jurisdictional fragmentation prevents global roll-out', 'Regulatory', 4, 3, 'Start in 2–3 aligned markets. Jurisdiction-specific API response profiles. Governance body includes multi-jurisdiction government seats.'),
    ('R4', 'Mandatory recall or liability issue with chip hardware', 'Regulatory', 2, 4, 'VSE-1 specifies automotive-grade chips with 15+ year design life. Liability framework established in governance model. Redundancy via three-node design.'),
]

for i, (ref, desc, cat, likelihood, impact, mitigation) in enumerate(risks):
    r = i + 5
    score = likelihood * impact
    if score >= 12:
        score_bg = DANGER_BG; score_fg = DANGER
    elif score >= 6:
        score_bg = WARN_BG; score_fg = WARN
    else:
        score_bg = SUCCESS_BG; score_fg = SUCCESS
    
    row_bg = LIGHT_ROW if i % 2 == 0 else LIGHT_ROW2
    
    ws3.cell(r, 1, ref).font = font(bold=True, color=PURPLE_MID, size=9)
    ws3.cell(r, 1).fill = fill(PURPLE_LIGHT)
    ws3.cell(r, 1).border = thin_border()
    ws3.cell(r, 1).alignment = center()
    
    ws3.cell(r, 2, desc).font = font(bold=False, color=DARK_TEXT)
    ws3.cell(r, 2).fill = fill(row_bg)
    ws3.cell(r, 2).border = thin_border()
    ws3.cell(r, 2).alignment = left()
    
    cat_colors = {'Technical': (PURPLE_LIGHT, PURPLE_MID), 'Commercial': (CYAN_LIGHT, CYAN_DARK), 'Regulatory': (WARN_BG, WARN)}
    cat_bg, cat_fg = cat_colors.get(cat, (LIGHT_ROW, DARK_TEXT))
    ws3.cell(r, 3, cat).font = font(bold=True, color=cat_fg, size=9)
    ws3.cell(r, 3).fill = fill(cat_bg)
    ws3.cell(r, 3).border = thin_border()
    ws3.cell(r, 3).alignment = center()
    
    ws3.cell(r, 4, likelihood).font = font(bold=True, color=score_fg)
    ws3.cell(r, 4).fill = fill(score_bg)
    ws3.cell(r, 4).border = thin_border()
    ws3.cell(r, 4).alignment = center()
    
    ws3.cell(r, 5, impact).font = font(bold=True, color=score_fg)
    ws3.cell(r, 5).fill = fill(score_bg)
    ws3.cell(r, 5).border = thin_border()
    ws3.cell(r, 5).alignment = center()
    
    ws3.cell(r, 6, mitigation).font = font(color=DARK_TEXT)
    ws3.cell(r, 6).fill = fill(row_bg)
    ws3.cell(r, 6).border = thin_border()
    ws3.cell(r, 6).alignment = left()
    ws3.row_dimensions[r].height = 40

spacer(ws3, len(risks) + 6)

# Legend
legend_row = len(risks) + 7
ws3.cell(legend_row, 1, 'LEGEND').font = font(bold=True, color=DARK_TEXT, size=9)
ws3.cell(legend_row, 1).fill = fill(LIGHT_ROW)
ws3.cell(legend_row, 1).border = thin_border()
ws3.cell(legend_row, 2, 'Score 12–25 = High Risk').font = font(bold=True, color=DANGER)
ws3.cell(legend_row, 2).fill = fill(DANGER_BG)
ws3.cell(legend_row, 2).border = thin_border()
ws3.cell(legend_row, 3, 'Score 6–11 = Medium Risk').font = font(bold=True, color=WARN)
ws3.cell(legend_row, 3).fill = fill(WARN_BG)
ws3.cell(legend_row, 3).border = thin_border()
ws3.cell(legend_row, 4, 'Score 1–5 = Low Risk').font = font(bold=True, color=SUCCESS)
ws3.cell(legend_row, 4).fill = fill(SUCCESS_BG)
ws3.cell(legend_row, 4).border = thin_border()
ws3.row_dimensions[legend_row].height = 22


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 4: STAKEHOLDER MAP
# ═══════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet('Stakeholder Access')
ws4.sheet_view.showGridLines = False

for col, w in [(1,28),(2,14),(3,14),(4,14),(5,14),(6,14),(7,14)]:
    set_col_width(ws4, col, w)

title_cell(ws4, 1, 1, 'Stakeholder Access Matrix — Who Sees What', 7, DARK_BG, 14)
subtitle_cell(ws4, 2, 1, '✓ = Full access  |  ~ = Partial / own data only  |  — = No access (enforced by ZK proof design)', 7, PURPLE_LIGHT, 9, PURPLE_MID)
spacer(ws4, 3)

header_row(ws4, 4, ['Data / Action', 'Manufacturer', 'Law\nEnforcement', 'Insurance', 'Dealer', 'Government\n/ DMV', 'Owner'], bg=DARK_BG, size=9)

access_data = [
    ('IDENTITY VERIFICATION', '', '', '', '', '', ''),
    ('Vehicle authentic? (Yes / No)', '✓', '✓', '✓', '✓', '✓', '✓'),
    ('Stolen / Flagged status', '✓', '✓  + Can set', '✓', '✓', '✓  + Can set', '✓'),
    ('Recall / Safety flags', '✓  + Can set', '✓', '✓', '✓', '✓', '✓'),
    ('OWNERSHIP DATA', '', '', '', '', '', ''),
    ('Ownership count (number of owners)', 'Count only', '—', '✓', '✓', '✓', 'Own only'),
    ('Full ownership history chain', '—', '—', '—', '—', '✓', 'Own only'),
    ('Previous owner personal details', '—', '—', '—', '—', '—', '—'),
    ('VEHICLE DATA', '', '', '', '', '', ''),
    ('VIN (plain text)', 'Own fleet', '—', '—', '—', '✓', 'Own only'),
    ('Engine / chassis serial numbers', 'Own fleet', '—', '—', '—', 'Via warrant', 'Own only'),
    ('Build specification', 'Full (own)', 'Basic', 'Basic', 'Basic', 'Basic', 'Basic'),
    ('HISTORY & EVENTS', '', '', '', '', '', ''),
    ('Service event count', '✓', '—', '✓', '✓', '—', '✓'),
    ('Specific service record details', '✓  (own)', '—', '—', '—', '—', '✓'),
    ('Insurance claim count', '—', '—', 'Own claims', '—', '—', 'Own only'),
    ('Insurance claim details / amounts', '—', '—', 'Own claims', '—', '—', 'Own only'),
    ('ACTIONS', '', '', '', '', '', ''),
    ('Mint Vehicle Identity Token', 'At manufacture', '—', '—', '—', '—', '—'),
    ('Transfer ownership / sell token', 'At manufacture', '—', '—', '✓  (new sale)', '—', '✓'),
    ('Flag vehicle as stolen', '—', '✓', '—', '—', '✓', 'Report only'),
    ('Log service event', '—', '—', '—', '—', '—', 'With consent'),
    ('Log insurance claim', '—', '—', 'Own claims', '—', '—', '—'),
]

for i, row_data in enumerate(access_data):
    r = i + 5
    is_section = row_data[1] == '' and row_data[0] != ''
    
    if is_section:
        for col in range(1, 8):
            c = ws4.cell(r, col, row_data[col-1] if col == 1 else '')
            c.fill = fill(PURPLE_MID)
            c.font = font(bold=True, color=WHITE, size=9)
            c.border = thin_border()
            c.alignment = left()
        ws4.row_dimensions[r].height = 20
    else:
        row_bg = LIGHT_ROW if i % 2 == 0 else LIGHT_ROW2
        ws4.cell(r, 1, row_data[0]).font = font(color=DARK_TEXT)
        ws4.cell(r, 1).fill = fill(row_bg)
        ws4.cell(r, 1).border = thin_border()
        ws4.cell(r, 1).alignment = left()
        
        for col in range(2, 8):
            val = row_data[col-1]
            if val == '✓' or 'Can set' in str(val):
                fg = SUCCESS; bg_c = SUCCESS_BG
            elif val == '—' or val == '':
                fg = GRAY_TEXT; bg_c = row_bg
            else:
                fg = CYAN_DARK; bg_c = CYAN_LIGHT
            
            c = ws4.cell(r, col, val)
            c.font = font(bold=(val == '✓'), color=fg, size=9)
            c.fill = fill(bg_c)
            c.border = thin_border()
            c.alignment = center()
        ws4.row_dimensions[r].height = 22


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 5: FRAMEWORK STANDARDS
# ═══════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet('Framework Standards')
ws5.sheet_view.showGridLines = False

for col, w in [(1,12),(2,22),(3,44),(4,16)]:
    set_col_width(ws5, col, w)

title_cell(ws5, 1, 1, 'Industry Framework Standards — Manufacturer Agreement Requirements', 4, DARK_BG, 14)
subtitle_cell(ws5, 2, 1, 'Four standards areas that all participating manufacturers must adopt. Governed by VGM-1 consortium.', 4, PURPLE_LIGHT, 9, PURPLE_MID)
spacer(ws5, 3)

framework = [
    ('VSE-1', 'Vehicle Secure Element Standard', '', 'HARDWARE'),
    ('VSE-1.1', 'Physical Specification', 'Automotive-grade (AEC-Q100 Grade 1). Operating temp -40°C to +125°C. IP67 minimum. Designed for 15+ year vehicle lifespan. Each chip must survive the manufacturing process (stamping, welding, painting, heat curing) with no degradation.', 'Required'),
    ('VSE-1.2', 'Cryptographic Specification', 'Threshold key generation scheme — 3-of-3. Elliptic curve cryptography minimum Ed25519 or stronger. Hardware random number generator (HRNG). Secure boot with key attestation. Physical Unclonable Function (PUF) binding to manufacture state. Anti-rollback firmware protection.', 'Required'),
    ('VSE-1.3', 'Communication Specification', 'CAN bus or Automotive Ethernet (AUTOSAR-compliant) for inter-node communication. Standardised command set for proof generation requests. Session-unique challenge-response protocol to prevent replay attacks. All inter-node traffic encrypted.', 'Required'),
    ('VSE-1.4', 'Tamper Response', 'Mandatory cryptographic key wipe on physical housing penetration. Alert broadcast via TN-3 on tamper detection in EN-1 or CN-2. All tamper events logged as immutable records on Midnight Network with timestamp and location data.', 'Required'),
    ('', '', '', ''),
    ('VIT-1', 'Vehicle Identity Token Standard', '', 'BLOCKCHAIN'),
    ('VIT-1.1', 'Public State Fields', 'Token ID (UUID v4), Manufacturer Code (ISO 3166-1 + OEM ID), VIN Hash (SHA-256 of plain VIN — not plain VIN itself), Token creation timestamp (UTC), Current status (ACTIVE / STOLEN / FLAGGED / EOL), VIT Standard version number.', 'Required'),
    ('VIT-1.2', 'Private / ZK-Accessible Fields', 'Plain VIN, Chassis number, Engine serial number, Build specification hash, Factory location code (ISO 3166-1), Production date. Stored in Midnight private state layer. Accessible only via authorised ZK proof by credentialed role.', 'Required'),
    ('VIT-1.3', 'Ownership Record Structure', 'Encrypted ownership history. Each transfer appends a new entry with timestamp and proof of authority. Previous owner entries remain accessible to current owner and government; access by others requires current owner ZK consent.', 'Required'),
    ('VIT-1.4', 'Event Log Types', 'MANUFACTURE — initial mint. SALE — ownership transfer. SERVICE — maintenance event (requires authorised service provider ZK proof). CLAIM — insurance event. FLAG_STOLEN — law enforcement only. FLAG_RECOVERED — law enforcement or government. FLAG_RECALL — manufacturer only. END_OF_LIFE — deregistration.', 'Required'),
    ('', '', '', ''),
    ('VAP-1', 'Verification API & Protocol Standard', '', 'API'),
    ('VAP-1.1', 'Query Protocol', 'REST and GraphQL API. Authentication via OAuth 2.0 with role-based access tokens issued by the VGM-1 governance body. Rate limiting enforced per role. All queries logged for audit trail with requesting party identity and timestamp.', 'Required'),
    ('VAP-1.2', 'Response Format', 'Standardised JSON schema with versioning. Status field always present. Data fields vary strictly by authenticated role. Response includes timestamp, proof hash for auditability. Maximum response time SLA: 30 seconds for live vehicle ZK proof on optimised SE hardware; under 1 second for cached status.', 'Required'),
    ('VAP-1.3', 'Offline Mode', 'Cached status certificates for areas with intermittent connectivity. Standard validity: 24 hours for general checks, 1 hour for stolen flag updates. Certificates cryptographically signed by Midnight Network validators.', 'Required'),
    ('', '', '', ''),
    ('VGM-1', 'Governance Model', '', 'GOVERNANCE'),
    ('VGM-1.1', 'Consortium Structure', 'Standards body with seats for: manufacturers (weighted by annual production volume), government representatives (minimum 3 jurisdictions), insurance industry, consumer advocacy, and Midnight Network as technical seat. Decisions require two-thirds supermajority.', 'Required'),
    ('VGM-1.2', 'Standard Change Process', 'Proposed changes published for 90-day public comment period. Technical review by engineering committee. Two-thirds vote required for adoption. Emergency security patches fast-tracked via executive committee with ratification within 30 days.', 'Required'),
    ('VGM-1.3', 'Dispute Resolution', 'Defined arbitration process for identity disputes. On-chain audit trail provides primary evidence. Binding decisions implemented via authorised smart contract calls. Appeals process through independent technical panel.', 'Required'),
    ('VGM-1.4', 'Access Revocation', 'Process for revoking API access for non-compliant parties. Manufacturer expulsion process for VSE-1 violations. Government access override protocol for emergency situations.', 'Required'),
]

for i, (code, name, detail, category) in enumerate(framework):
    r = i + 4
    is_header = name == '' and code != ''
    is_spacer = code == '' and name == ''
    
    if is_spacer:
        ws5.row_dimensions[r].height = 8
        continue
    
    if is_header:
        for col in range(1, 5):
            ws5.cell(r, col).fill = fill(DARK_BG)
        ws5.cell(r, 1, code).font = font(bold=True, color=WHITE, size=11)
        ws5.cell(r, 1).fill = fill(PURPLE_MID)
        ws5.cell(r, 2, name).font = font(bold=True, color=WHITE, size=11)
        ws5.cell(r, 2).fill = fill(PURPLE_MID)
        ws5.cell(r, 3, '').fill = fill(DARK_BG)
        cat_colors_map = {'HARDWARE': (WARN_BG, WARN), 'BLOCKCHAIN': (CYAN_LIGHT, CYAN_DARK), 'API': (PURPLE_LIGHT, PURPLE_MID), 'GOVERNANCE': (SUCCESS_BG, SUCCESS)}
        cb, cf = cat_colors_map.get(category, (DARK_BG, WHITE))
        ws5.cell(r, 4, category).font = font(bold=True, color=cf, size=9)
        ws5.cell(r, 4).fill = fill(cb)
        ws5.cell(r, 4).alignment = center()
        ws5.row_dimensions[r].height = 26
    else:
        row_bg = LIGHT_ROW if i % 2 == 0 else LIGHT_ROW2
        ws5.cell(r, 1, code).font = font(bold=True, color=PURPLE_MID, size=9)
        ws5.cell(r, 1).fill = fill(PURPLE_LIGHT)
        ws5.cell(r, 1).border = thin_border()
        ws5.cell(r, 1).alignment = center()
        
        ws5.cell(r, 2, name).font = font(bold=True, color=DARK_TEXT, size=9)
        ws5.cell(r, 2).fill = fill(row_bg)
        ws5.cell(r, 2).border = thin_border()
        ws5.cell(r, 2).alignment = left()
        
        ws5.cell(r, 3, detail).font = font(color=DARK_TEXT, size=9)
        ws5.cell(r, 3).fill = fill(row_bg)
        ws5.cell(r, 3).border = thin_border()
        ws5.cell(r, 3).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        ws5.cell(r, 4, category if category else 'Required').font = font(bold=True, color=SUCCESS, size=9)
        ws5.cell(r, 4).fill = fill(SUCCESS_BG if category else row_bg)
        ws5.cell(r, 4).border = thin_border()
        ws5.cell(r, 4).alignment = center()
        ws5.row_dimensions[r].height = 60

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 6: SCENARIOS
# ═══════════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet('Scenarios')
ws6.sheet_view.showGridLines = False

for col, w in [(1,32),(2,16),(3,16),(4,16),(5,22)]:
    set_col_width(ws6, col, w)

title_cell(ws6, 1, 1, 'ShieldVIN — Scenario Analysis', 5, DARK_BG, 14)
subtitle_cell(ws6, 2, 1, 'Conservative / Base / Aggressive — switchable assumptions  |  All revenue in USD Millions', 5, PURPLE_LIGHT, 10, PURPLE_MID)
spacer(ws6, 3)

# ── ASSUMPTION INPUTS ──
subtitle_cell(ws6, 4, 1, '  SCENARIO ASSUMPTIONS  (blue cells are inputs)', 5, DARK_BG, 10, WHITE)
ws6.row_dimensions[4].height = 22

header_row(ws6, 5, ['Assumption', 'Conservative', 'Base Case', 'Aggressive', 'Notes'], bg=PURPLE_MID)

BLUE_INPUT  = '1E3A8A'
BLUE_BG     = 'DBEAFE'

assumptions = [
    ('Year 5 enrolled vehicles (M units)',         0.8,   2.0,   5.0,   'Global new vehicle production ~80M/yr. Conservative = 1%, Base = 2.5%, Aggressive = 6.25%'),
    ('Blended minting fee per vehicle (USD)',       8.00,  12.00, 18.00, 'Volume tier pricing. Low = entry/loss-leader, High = premium OEM tier'),
    ('API queries per enrolled vehicle per year',  3.0,   6.0,   10.0,  'Avg queries across insurer, police, dealer, owner channels'),
    ('Blended API fee per query (USD)',            0.12,  0.22,  0.40,  'Mix of $0.10 police/govt, $0.25 insurer, $2.50 enterprise'),
    ('Enterprise partnership revenue Year 5 (USD M)', 8.0, 19.0, 42.0, 'Revenue share from insurance claim savings + data products'),
    ('DUST fee per transaction (USD equiv.)',      0.01,  0.03,  0.08,  'Operational cost. NIGHT/DUST market rate sensitivity. Not a revenue item.'),
    ('Year 5 DUST cost (USD M, total fleet)',      0.30,  1.50,  8.00,  'Estimated total on-chain activity cost. Sensitive to NIGHT price.'),
]

for i, (label, cons, base, aggr, note) in enumerate(assumptions):
    r = i + 6
    row_bg = LIGHT_ROW if i % 2 == 0 else LIGHT_ROW2

    ws6.cell(r, 1, label).font = font(color=DARK_TEXT)
    ws6.cell(r, 1).fill = fill(row_bg)
    ws6.cell(r, 1).border = thin_border()
    ws6.cell(r, 1).alignment = left()

    for j, val in [(2, cons), (3, base), (4, aggr)]:
        c = ws6.cell(r, j, val)
        c.font = font(bold=True, color=BLUE_INPUT)
        c.fill = fill(BLUE_BG)
        c.border = thin_border()
        c.alignment = center()

    ws6.cell(r, 5, note).font = font(color=GRAY_TEXT, italic=True, size=9)
    ws6.cell(r, 5).fill = fill(row_bg)
    ws6.cell(r, 5).border = thin_border()
    ws6.cell(r, 5).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws6.row_dimensions[r].height = 32

# ── CALCULATED OUTPUTS ──
spacer(ws6, 14)
subtitle_cell(ws6, 15, 1, '  YEAR 5 CALCULATED OUTPUTS  (formulas reference assumption cells above)', 5, DARK_BG, 10, WHITE)
ws6.row_dimensions[15].height = 22
header_row(ws6, 16, ['Output', 'Conservative', 'Base Case', 'Aggressive', 'Formula Basis'], bg=PURPLE_MID)

# Row references: assumptions start at row 6
# B6=enrolled vehicles (M), B7=minting fee, B8=queries/vehicle, B9=API fee, B10=enterprise, B11=DUST/tx, B12=DUST total cost
# Columns: B=Conservative, C=Base, D=Aggressive

outputs = [
    ('Minting Revenue Year 5 (USD M)',
     '=B6*1000000*B7/1000000', '=C6*1000000*C7/1000000', '=D6*1000000*D7/1000000',
     'Enrolled vehicles × minting fee per vehicle'),
    ('API Revenue Year 5 (USD M)',
     '=B6*1000000*B8*B9/1000000', '=C6*1000000*C8*C9/1000000', '=D6*1000000*D8*D9/1000000',
     'Enrolled vehicles × queries/vehicle × fee/query'),
    ('Enterprise Revenue Year 5 (USD M)',
     '=B10', '=C10', '=D10',
     'Direct input — partnership revenue'),
    ('Total Revenue Year 5 (USD M)',
     '=B17+B18+B19', '=C17+C18+C19', '=D17+D18+D19',
     'Sum of all three streams'),
    ('DUST Operational Cost Year 5 (USD M)',
     '=B12', '=C12', '=D12',
     'Estimated on-chain activity cost — not revenue'),
    ('Net Revenue after DUST Cost (USD M)',
     '=B20-B21', '=C20-C21', '=D20-D21',
     'Total Revenue minus DUST fees'),
    ('Revenue per Enrolled Vehicle (USD)',
     '=IF(B6>0,(B20/B6),0)', '=IF(C6>0,(C20/C6),0)', '=IF(D6>0,(D20/D6),0)',
     'Net revenue per enrolled vehicle — sanity check vs $25 blended target'),
]

for i, (label, cons_f, base_f, aggr_f, note) in enumerate(outputs):
    r = i + 17
    is_total = 'Total Revenue' in label or 'Net Revenue' in label
    row_bg = DARK_BG if is_total else (LIGHT_ROW if i % 2 == 0 else LIGHT_ROW2)
    fg = WHITE if is_total else DARK_TEXT
    val_color = WHITE if is_total else PURPLE_MID

    ws6.cell(r, 1, label).font = font(bold=is_total, color=fg)
    ws6.cell(r, 1).fill = fill(row_bg)
    ws6.cell(r, 1).border = thin_border()
    ws6.cell(r, 1).alignment = left()

    for j, formula in [(2, cons_f), (3, base_f), (4, aggr_f)]:
        c = ws6.cell(r, j, formula)
        c.font = font(bold=is_total, color=val_color)
        c.fill = fill(row_bg)
        c.border = thin_border()
        c.alignment = center()
        c.number_format = '#,##0.00'

    ws6.cell(r, 5, note).font = font(color=GRAY_TEXT if not is_total else PURPLE_LIGHT, italic=True, size=9)
    ws6.cell(r, 5).fill = fill(row_bg)
    ws6.cell(r, 5).border = thin_border()
    ws6.cell(r, 5).alignment = left()
    ws6.row_dimensions[r].height = 22

# ── DUST SENSITIVITY NOTE ──
spacer(ws6, 25)
subtitle_cell(ws6, 26, 1, '  DUST TOKEN ECONOMICS — SENSITIVITY NOTE', 5, DARK_BG, 10, WHITE)
ws6.row_dimensions[26].height = 22

dust_notes = [
    ('DUST is Midnight Network\'s transaction fee token. Every vehicle mint, ownership transfer, verification query, and lifecycle event consumes DUST.',
     'Operational cost — not revenue for ShieldVIN'),
    ('DUST is purchased by burning NIGHT (Midnight\'s staking/governance token). The NIGHT/DUST conversion rate is set by the network, not by ShieldVIN.',
     'Price risk: NIGHT market price drives DUST cost'),
    ('At ~0.03 DUST per transaction (Midnight Network testnet data, early 2026), and assuming ~5 on-chain events per enrolled vehicle per year, the Base Case fleet of 2M vehicles generates ~300M DUST transactions/year.',
     '~300M DUST/yr at 2M enrolled vehicles (Base)'),
    ('DUST cost sensitivity: if NIGHT appreciates 10×, the DUST operational cost (in USD) increases proportionally. The model assumes mid-range NIGHT pricing. See row 12 above for sensitivity input.',
     'Hedge strategy: pre-purchase DUST in bulk at manufacture'),
    ('The high on-chain activity is a feature for the Midnight Network pitch: ShieldVIN becomes one of the largest institutional DUST consumers on the network, supporting the token economy.',
     'Strategic value to Midnight — not just a cost'),
]

header_row(ws6, 27, ['Note', 'Implication'], bg=PURPLE_MID)
ws6.merge_cells('A27:C27')
ws6.cell(27, 4).fill = ws6.cell(27, 1).fill
ws6.cell(27, 4).font = ws6.cell(27, 1).font
ws6.cell(27, 4).border = thin_border()
ws6.cell(27, 5).fill = ws6.cell(27, 1).fill
ws6.cell(27, 5).border = thin_border()

for i, (note_text, implication) in enumerate(dust_notes):
    r = i + 28
    row_bg = LIGHT_ROW if i % 2 == 0 else LIGHT_ROW2
    ws6.cell(r, 1, note_text).font = font(color=DARK_TEXT, size=9)
    ws6.merge_cells(f'A{r}:C{r}')
    ws6.cell(r, 1).fill = fill(row_bg)
    ws6.cell(r, 1).border = thin_border()
    ws6.cell(r, 1).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws6.cell(r, 4, implication).font = font(bold=True, color=PURPLE_MID, size=9)
    ws6.cell(r, 4).fill = fill(row_bg)
    ws6.cell(r, 4).border = thin_border()
    ws6.cell(r, 4).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws6.merge_cells(f'D{r}:E{r}')
    ws6.row_dimensions[r].height = 42

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 7: REVENUE CHART
# ═══════════════════════════════════════════════════════════════════════════
ws7 = wb.create_sheet('Revenue Chart')
ws7.sheet_view.showGridLines = False

for col, w in [(1,30),(2,14),(3,14),(4,14),(5,14),(6,14),(7,16)]:
    set_col_width(ws7, col, w)

title_cell(ws7, 1, 1, 'Revenue Build — 5-Year Stacked Waterfall', 7, DARK_BG, 14)
subtitle_cell(ws7, 2, 1, 'Data sourced from Revenue Model sheet. Chart updates automatically when assumptions change.', 7, PURPLE_LIGHT, 9, PURPLE_MID)
spacer(ws7, 3)

# Data table header
header_row(ws7, 4, ['Revenue Stream', 'Year 1\n2026', 'Year 2\n2027', 'Year 3\n2028', 'Year 4\n2029', 'Year 5\n2030', 'Notes'], bg=PURPLE_MID)

streams_chart = [
    ('Vehicle Identity Token Minting',
     "='Revenue Model'!B34", "='Revenue Model'!C34", "='Revenue Model'!D34", "='Revenue Model'!E34", "='Revenue Model'!F34",
     'Per-vehicle fee at manufacture'),
    ('Verification API Access',
     "='Revenue Model'!B35", "='Revenue Model'!C35", "='Revenue Model'!D35", "='Revenue Model'!E35", "='Revenue Model'!F35",
     'Insurance, dealers, fleet operators'),
    ('Insurance & Enterprise',
     "='Revenue Model'!B36", "='Revenue Model'!C36", "='Revenue Model'!D36", "='Revenue Model'!E36", "='Revenue Model'!F36",
     'Partnership & data product revenue'),
    ('TOTAL REVENUE',
     '=B5+B6+B7', '=C5+C6+C7', '=D5+D6+D7', '=E5+E6+E7', '=F5+F6+F7',
     'All three streams combined'),
]

for i, (stream, y1, y2, y3, y4, y5, note) in enumerate(streams_chart):
    r = i + 5
    is_total = 'TOTAL' in stream
    bg_color = DARK_BG if is_total else (LIGHT_ROW if i % 2 == 0 else LIGHT_ROW2)
    fg_color = WHITE if is_total else DARK_TEXT

    ws7.cell(r, 1, stream).font = font(bold=is_total, color=fg_color)
    ws7.cell(r, 1).fill = fill(bg_color)
    ws7.cell(r, 1).border = thin_border()
    ws7.cell(r, 1).alignment = left()

    for j, val in enumerate([y1, y2, y3, y4, y5], 2):
        c = ws7.cell(r, j, val)
        c.font = font(bold=is_total, color=WHITE if is_total else PURPLE_MID)
        c.fill = fill(bg_color)
        c.border = thin_border()
        c.alignment = center()
        c.number_format = '$#,##0.00"M"'

    ws7.cell(r, 7, note).font = font(color=GRAY_TEXT if not is_total else WHITE, italic=True, size=9)
    ws7.cell(r, 7).fill = fill(bg_color)
    ws7.cell(r, 7).border = thin_border()
    ws7.cell(r, 7).alignment = left()
    ws7.row_dimensions[r].height = 22 if not is_total else 26

spacer(ws7, 10)

# ── STACKED BAR CHART ──
chart = BarChart()
chart.type = "col"
chart.grouping = "stacked"
chart.overlap = 100
chart.title = "ShieldVIN — Revenue Build 2026–2030 (USD Millions)"
chart.style = 10
chart.y_axis.title = "Revenue (USD Millions)"
chart.x_axis.title = "Year"
chart.width = 22
chart.height = 14

# Each row (5-7) is one revenue stream; cols B-F are years 1-5
# Col A of each row is the series title
data_with_titles = Reference(ws7, min_col=1, max_col=6, min_row=5, max_row=7)
chart.add_data(data_with_titles, from_rows=True, titles_from_data=True)

# Year categories from row 4, cols B-F
cats = Reference(ws7, min_col=2, max_col=6, min_row=4)
chart.set_categories(cats)

ws7.add_chart(chart, "A12")


# ═══════════════════════════════════════════════════════════════════════════
# STAKEHOLDER ACCESS — CONDITIONAL FORMATTING (on top of static colours)
# ═══════════════════════════════════════════════════════════════════════════
# access_data starts at row 5; 22 rows → data range B5:G26
CF_RANGE = 'B5:G26'

# Green highlight for any cell containing the ✓ character
green_fill_cf  = PatternFill(fill_type='solid', fgColor=SUCCESS_BG)
green_font_cf  = Font(name='Arial', bold=True, color=SUCCESS, size=9)
ws4.conditional_formatting.add(CF_RANGE, FormulaRule(
    formula=['NOT(ISERROR(SEARCH("✓",B5)))'],
    fill=green_fill_cf,
    font=green_font_cf,
))

# Cyan / partial for cells that are not ✓, not blank, and not —
partial_fill_cf = PatternFill(fill_type='solid', fgColor=CYAN_LIGHT)
partial_font_cf = Font(name='Arial', color=CYAN_DARK, size=9)
ws4.conditional_formatting.add(CF_RANGE, FormulaRule(
    formula=['AND(B5<>"—",B5<>"",ISERROR(SEARCH("✓",B5)))'],
    fill=partial_fill_cf,
    font=partial_font_cf,
))

# Dim gray for — (no access) cells
no_access_fill = PatternFill(fill_type='solid', fgColor='F0F0F0')
no_access_font = Font(name='Arial', color='AAAAAA', size=9)
ws4.conditional_formatting.add(CF_RANGE, FormulaRule(
    formula=['B5="—"'],
    fill=no_access_fill,
    font=no_access_font,
))


# ═══════════════════════════════════════════════════════════════════════════
# DATA VALIDATION — REVENUE MODEL (ws2)
# ═══════════════════════════════════════════════════════════════════════════

# Row 8: % of production enrolled — decimal 0 to 1
dv_pct_enroll = DataValidation(
    type='decimal', operator='between', formula1=0, formula2=1,
    allow_blank=True,
    showErrorMessage=True,
    errorTitle='Invalid Input',
    error='Enter a decimal between 0.0 and 1.0 (e.g. 0.05 = 5%)',
    showInputMessage=True,
    promptTitle='% Enrolled',
    prompt='Decimal between 0.0 (0%) and 1.0 (100%)',
)
dv_pct_enroll.sqref = 'B8:F8'
ws2.add_data_validation(dv_pct_enroll)

# Row 10: Average fee per token — integer 1 to 50
dv_mint_fee = DataValidation(
    type='decimal', operator='between', formula1=1, formula2=50,
    allow_blank=True,
    showErrorMessage=True,
    errorTitle='Invalid Input',
    error='Minting fee must be between $1 and $50 per vehicle',
    showInputMessage=True,
    promptTitle='Minting Fee (USD)',
    prompt='Enter a value between 1 and 50 (USD per vehicle)',
)
dv_mint_fee.sqref = 'B10:F10'
ws2.add_data_validation(dv_mint_fee)

# Row 16: Avg verifications per vehicle per year — integer 1 to 50
dv_queries = DataValidation(
    type='whole', operator='between', formula1=1, formula2=50,
    allow_blank=True,
    showErrorMessage=True,
    errorTitle='Invalid Input',
    error='Queries per vehicle should be between 1 and 50',
    showInputMessage=True,
    promptTitle='Queries per Vehicle',
    prompt='Expected annual queries per enrolled vehicle (1–50)',
)
dv_queries.sqref = 'B16:F16'
ws2.add_data_validation(dv_queries)

# Row 18: % queries paid — decimal 0 to 1
dv_pct_paid = DataValidation(
    type='decimal', operator='between', formula1=0, formula2=1,
    allow_blank=True,
    showErrorMessage=True,
    errorTitle='Invalid Input',
    error='Enter a decimal between 0.0 and 1.0',
    showInputMessage=True,
    promptTitle='% Paid Queries',
    prompt='Fraction of total queries that are paid (0.0–1.0)',
)
dv_pct_paid.sqref = 'B18:F18'
ws2.add_data_validation(dv_pct_paid)

# Row 19: Average price per query — decimal 0.01 to 10
dv_query_price = DataValidation(
    type='decimal', operator='between', formula1=0.01, formula2=10,
    allow_blank=True,
    showErrorMessage=True,
    errorTitle='Invalid Input',
    error='Average query price must be between $0.01 and $10.00',
    showInputMessage=True,
    promptTitle='Avg Query Price (USD)',
    prompt='Blended average price per paid verification query ($0.01–$10)',
)
dv_query_price.sqref = 'B19:F19'
ws2.add_data_validation(dv_query_price)


# ═══════════════════════════════════════════════════════════════════════════
# DATA VALIDATION — SCENARIOS (ws6)
# ═══════════════════════════════════════════════════════════════════════════

# Row 6: Year 5 enrolled vehicles (M) — 0.01 to 80
dv_enrolled_m = DataValidation(
    type='decimal', operator='between', formula1=0.01, formula2=80,
    allow_blank=True,
    showErrorMessage=True,
    errorTitle='Invalid Input',
    error='Enrolled vehicles must be between 0.01M and 80M (global production cap)',
    showInputMessage=True,
    promptTitle='Enrolled Vehicles (M)',
    prompt='Year 5 enrolled vehicles in millions (0.01–80)',
)
dv_enrolled_m.sqref = 'B6:D6'
ws6.add_data_validation(dv_enrolled_m)

# Row 7: Minting fee — 1 to 50
dv_sc_mint = DataValidation(
    type='decimal', operator='between', formula1=1, formula2=50,
    allow_blank=True,
    showErrorMessage=True,
    errorTitle='Invalid Input',
    error='Blended minting fee must be between $1 and $50',
    showInputMessage=True,
    promptTitle='Minting Fee (USD)',
    prompt='Blended minting fee per vehicle in USD (1–50)',
)
dv_sc_mint.sqref = 'B7:D7'
ws6.add_data_validation(dv_sc_mint)

# Row 8: API queries per vehicle — 1 to 50
dv_sc_queries = DataValidation(
    type='decimal', operator='between', formula1=1, formula2=50,
    allow_blank=True,
    showErrorMessage=True,
    errorTitle='Invalid Input',
    error='API queries per vehicle must be between 1 and 50',
    showInputMessage=True,
    promptTitle='Queries per Vehicle',
    prompt='Annual API queries per enrolled vehicle (1–50)',
)
dv_sc_queries.sqref = 'B8:D8'
ws6.add_data_validation(dv_sc_queries)

# Row 9: API fee per query — 0.01 to 10
dv_sc_api_fee = DataValidation(
    type='decimal', operator='between', formula1=0.01, formula2=10,
    allow_blank=True,
    showErrorMessage=True,
    errorTitle='Invalid Input',
    error='API fee per query must be between $0.01 and $10.00',
    showInputMessage=True,
    promptTitle='API Fee per Query (USD)',
    prompt='Blended average fee per verification query ($0.01–$10)',
)
dv_sc_api_fee.sqref = 'B9:D9'
ws6.add_data_validation(dv_sc_api_fee)


# Save
wb.save(os.path.join(os.path.dirname(os.path.abspath(__file__)), '../../dist/ShieldVIN_Financial_Model.xlsx'))
print('Excel created successfully')
